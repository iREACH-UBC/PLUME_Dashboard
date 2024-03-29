"""
Created by Julian Fawkes, 2020, and Chris Kelly, 2021/2022
Contributions and minor edits by: Naomi Zimmerman, Melanie MacArthur, Stefan Colbow, Rachel Habermehl, and Davi Monticelli
PLUME Dashboard - A browser-based pollutant data visualization program built using Dash. The dashboard pulls data
from a DAQ script using a locally-hosted redis server. Each instrument sends data using a
different transfer protocol and is taken in by a CR1000X datalogger.

Either modbus-tcp_daq.py or another DAQ script must first be running in order for this script to work. Additionally, the
'user_defined_settings.ini' file must first be set up before running this script.
"""

import os
import math
import sys
import dash
import csv
import statistics
from dash.dependencies import Output, Input, State
from plotly.subplots import make_subplots
from collections import deque
import collections
import plotly.express as px
import dash_core_components as dcc
import dash_html_components as html
import dash_bootstrap_components as dbc
import datetime as dt
import dash_daq as daq
import redis
import openpyxl
import pandas as pd
import numpy as np
from configparser import ConfigParser
from pathlib import Path
import json
from os.path import exists
import pyuac

#NOTE: when you see variables like "fold_start = 0" or "avs = 0", those are just lines of folds starting and ending (pycharm doesn't want fold borders to be comments)

'''#########################
## SECTION 1. Bookkeeping ##
#########################'''
fold_start = 0

# Create a dash object 'app' with name '__name__'. Stop the title from changing when content updates.
app = dash.Dash(__name__,update_title=None, external_stylesheets=[dbc.themes.MINTY])
server = app.server
app_color = dict(graph_bg="#E3E4E0", red="#E94029", green="#00703D", yellow="#E4BC3F")
app.title = 'PLUME Dashboard'


interval_s = 2 #Instrument polling rates in seconds.normally 2
graph_range = 2 * 60 #Range of graph display in seconds.
#trace_length = math.trunc(graph_range / interval_s)
trace_length = 60 #no. of points on the graph

#list of dropdown options
instrumentdict = [
    {"label": "NO2", "value": "NO2"},
    {"label": "WCPC", "value": "WCPC"},
    {"label": "O3", "value": "O3"},
    {"label": "CO", "value": "CO"},
    {"label": "CO2", "value": "CO2"},
    {"label": "NO", "value": "NO"},
    {"label": "WS", "value": "WS"},
    {"label": "WD", "value": "WD"},
]

#unit dictionary, this MUST be matched with instrumentdict
labeldict = {"NO2": "Concentration (ppb)", "WCPC": "Concentration (#/cm\u00B3)",
             "O3": "Concentration (ppb)", "CO": "Concentration (ppm)",
             "CO2": "Concentration (ppm)", "NO": "Concentration (ppb)",
             "WS": "Wind-speed (m/s)", "WD": "Wind-direction (degrees)"}

#define our deques
no2_trace_y = deque([0], maxlen=trace_length)
no2_trace_x = deque([0], maxlen=trace_length)
wcpc_trace_y = deque([0], maxlen=trace_length)
wcpc_trace_x = deque([0], maxlen=trace_length)
o3_trace_y = deque([0], maxlen=trace_length)
o3_trace_x = deque([0], maxlen=trace_length)
co_trace_y = deque([0], maxlen=trace_length)
co_trace_x = deque([0], maxlen=trace_length)
co2_trace_y = deque([0], maxlen=trace_length)
co2_trace_x = deque([0], maxlen=trace_length)
no_trace_y = deque([0], maxlen=trace_length)
no_trace_x = deque([0], maxlen=trace_length)
ws_trace_y = deque([0], maxlen=trace_length)
ws_trace_x = deque([0], maxlen=trace_length)
wd_trace_y = deque([0], maxlen=trace_length)
wd_trace_x = deque([0], maxlen=trace_length)

#define our container deques, these are used to pass data to the live plot
no2_trace_container = deque([dict(x=0, y=0)], maxlen=1)
wcpc_trace_container = deque([dict(x=0, y=0)], maxlen=1)
o3_trace_container = deque([dict(x=0, y=0)], maxlen=1)
co_trace_container = deque([dict(x=0, y=0)], maxlen=1)
co2_trace_container = deque([dict(x=0, y=0)], maxlen=1)
no_trace_container = deque([dict(x=0, y=0)], maxlen=1)
ws_trace_container = deque([dict(x=0, y=0)], maxlen=1)
wd_trace_container = deque([dict(x=0, y=0)], maxlen=1)

#helper function for settings loading, changes "int,int" into [int,int]
def string_to_list_interval(string_in):

    # remove all spaces from string_in
    string_in = string_in.replace(" ", "")

    #define vars
    comma_index = 0
    lower=""
    upper=""
    result=[]

    #append to lower until comma is found
    for i in range(0,len(string_in)):
        if string_in[i] == ",":
            comma_index=i
            break
        else:
            lower+=string_in[i]

    #append to upper until end of string, starting from first index after comma
    for i in range(comma_index+1,len(string_in)):
        upper+=string_in[i]

    #return result as list
    result.append(int(lower))
    result.append(int(upper))
    return result

foldend = 0


'''###################################
## SECTION 2. Auto event algorithms ##
###################################'''
avs = 0

#peak detection algorithm
def A1ap(data_points, pollutant):
    #exiting the function if A1 is disabled
    global A1
    global A1_startup_bypass
    global index_clock
    if (A1 == False) or (index_clock<A1_startup_bypass):
        return None

    # defining vars and importing global vars
    #m = statistics.median(data_points) #old command using 50th percentile (median)
    global A1_percentile
    global A1_thresh_bump_percentile
    m = np.percentile(data_points, A1_percentile[pollutant]) #using a custom percentile as our median
    below_m = []
    global A1_auto_event_count
    global A1_n
    global A1_coeff

    if A1_thresh_bump_percentile[pollutant] == 0:
        thresh_bump = 0
    else:
        thresh_bump = np.percentile(data_points, A1_thresh_bump_percentile[pollutant])

    #creating list of points below median and determining the threshold
    for i in data_points:
        if i < m:
            below_m.append(i)
    if len(below_m) < 2:
        return None
    sd = statistics.stdev(below_m) #stdev will do sample sd and pstdev will do population sd
    thresh = A1_coeff[pollutant] * sd
    thresh += thresh_bump #adding thresh bumpp

    #(printing info for debugging purposes, only implemented for NO2)
    '''
    if pollutant == "no2":
        global no2_clock_y
        global no2_clock_x
        below_m_rounded = [round(num, 2) for num in below_m]
        print("x="+str(no2_clock_x)+", y=" + str(round(no2_clock_y, 2)) + ", median: " + str(round(m, 2)) + ", bm-sd: " + str(
            round(sd, 2)) + ", thresh: " + str(round(thresh, 2)) + " A1_n=" + str(A1_n["no2"]) + ", below_m: " + str(
            below_m_rounded))
    '''

    # checking the appropriate threshold
    if A1_n[pollutant] == 0:
        if data_points[-1] >= thresh:
            A1_n[pollutant] += 1
            auto_event_mark("A1-"+pollutant+"-" + str(A1_auto_event_count[pollutant]),"peak",pollutant)
            A1_auto_event_count[pollutant] += 1

            #for debugging
            #if pollutant == "no2":
                #print("peak")

            return None
        else:
            return None
    else:
        if data_points[-1] >= (thresh + sd * math.sqrt(A1_n[pollutant])):
            A1_n[pollutant] += 1
            auto_event_mark("A1-" + pollutant + "-" + str(A1_auto_event_count[pollutant]),"peak",pollutant)
            A1_auto_event_count[pollutant] += 1

            # for debugging
            #if pollutant == "no2":
                #print("s-peak")

            return None
        else:
            A1_n[pollutant] = 0
            return None

#A2 was originally designed to detect a steady increase... however we have disabled it. The code is here for anyone who wants to dabble with it
def A2ap(data_points, pollutant):
    # exiting the function if A2 is disabled
    global A2
    if A2 == False:
        return None

    #importing global vars and defining local vars
    global A2_interval
    global A2_slope_thresh
    global A2_hits_to_sink
    global A2_n
    global A2_auto_event_count
    interval_size = A2_interval[pollutant].maxlen

    #importing next data point to the interval and exiting function if there isn't enough data
    A2_interval[pollutant].append(data_points[-1])
    if len(A2_interval[pollutant]) != interval_size:
        return None

    #computing slope and comparing to threshold
    slope = ( A2_interval[pollutant][-1] - A2_interval[pollutant][0] ) / interval_size
    if abs(slope) >= A2_slope_thresh[pollutant]:
        A2_n[pollutant] += 1
    else:
        A2_n[pollutant] = 0

    #debugging purpose, only implemented for no2
    if pollutant == "no2":
        rounded_interval = [round(num, 2) for num in A2_interval[pollutant]]
        print("slope="+str(round(slope,2))+", n="+str(A2_n[pollutant])+", interval: "+str(rounded_interval))
        if abs(slope) >= A2_slope_thresh[pollutant]:
            print("^HIT")

    if A2_n[pollutant] == A2_hits_to_sink[pollutant]:
        A2_n[pollutant] = 0

        #printing for debugging, only implemented for no2
        if pollutant == "no2":
            print("and sunk!")

        if slope > 0:
            auto_event_mark("A2-" + pollutant + "-I-" + str(A2_auto_event_count[pollutant]),"increase",pollutant)
            A2_auto_event_count[pollutant] += 1
            return None
        else:
            auto_event_mark("A2-" + pollutant + "-D-" + str(A2_auto_event_count[pollutant]),"decrease",pollutant)
            A2_auto_event_count[pollutant] += 1
            return None
    else:
        return None

#AQ over/under detection
def AQap(data_points, pollutant):
    #exitting function if AQ is disabled
    global AQ
    if AQ == False:
        return None

    #import global vars
    global AQ_thresh
    global AQ_over
    global AQ_auto_event_count

    #detecting if the graph moves upward over the line
    if (data_points[-1] > AQ_thresh[pollutant]) and (AQ_over[pollutant] == False):
        AQ_over[pollutant] = True
        #print(pollutant + " AQ over threshold of " + str(AQ_thresh[pollutant]))
        ##
        auto_event_mark("AQ-" + pollutant.upper() + "-over-" + str(AQ_auto_event_count[pollutant]), "AQ over", pollutant)
        AQ_auto_event_count[pollutant] += 1
        return None

    #detecting if the graph moves downward over the line
    if (data_points[-1] < AQ_thresh[pollutant]) and (AQ_over[pollutant] == True):
        AQ_over[pollutant] = False
        #print(pollutant + " AQ back to under threshold of " + str(AQ_thresh[pollutant]))
        auto_event_mark("AQ-" + pollutant.upper() + "-under-" + str(AQ_auto_event_count[pollutant]), "AQ under", pollutant)
        AQ_auto_event_count[pollutant] += 1
        return None

#detects when the wind direction is within a certain radial range
def wind_direction_alert(data_points, pollutant):
    #exitting function if wind direction alert is disabled
    global enable_wind_direction_alert
    if enable_wind_direction_alert == False:
        return None


    #import global vars, uses same counter variables as AQ algorithm
    global wind_direction_alert_range
    global AQ_over
    global AQ_auto_event_count

    #detecting if the direction moves INTO the alert range
    if (data_points[-1] >= wind_direction_alert_range[0]) and (data_points[-1] <= wind_direction_alert_range[1]) and (AQ_over[pollutant] == False):
        AQ_over[pollutant] = True
        print("Wind direction is within alert range")
        auto_event_mark("WD-alert-begin-" + str(AQ_auto_event_count[pollutant]), "WD alert begin", pollutant)
        AQ_auto_event_count[pollutant] += 1
        return None

    #detecting if the direction moves OUT OF the alert range
    if (not ( (data_points[-1] >= wind_direction_alert_range[0]) and (data_points[-1] <= wind_direction_alert_range[1]))) and (AQ_over[pollutant] == True):
        AQ_over[pollutant] = False
        print("Wind direction is no longer within alert range")
        auto_event_mark("WD-alert-end-" + str(AQ_auto_event_count[pollutant]), "WD alert end", pollutant)
        AQ_auto_event_count[pollutant] += 1
        return None

#helper function to remove the leftmost zero of traces so that A1 can work properly
def zero_flush():
    #import our global switch and exit function if it's already been ran or if the queues are too small
    global left_zero_popped
    global no2_trace_y
    if (left_zero_popped == True) or len(no2_trace_y) < 4:
        return None

    #import all of our other pollutant queues
    global wcpc_trace_y
    global o3_trace_y
    global co_trace_y
    global co2_trace_y
    global no_trace_y
    global ws_trace_y
    global wd_trace_y

    #delete the leftmost entry (the trailing 0)
    #print("flushing zeros")
    no2_trace_y.popleft()
    wcpc_trace_y.popleft()
    o3_trace_y.popleft()
    co_trace_y.popleft()
    co2_trace_y.popleft()
    no_trace_y.popleft()
    ws_trace_y.popleft()
    wd_trace_y.popleft()

    #prevent function from running again
    left_zero_popped = True


srgdfg = 0


'''##################
## SECTION 3. Dash ##
##################'''
avs=0
# Helper function definitions for complex or repeated operations.
def create_graduatedbar_helper(name):
    """Helper function to create graduated bars (saves lines of code)"""

    # Create each graduated bar row
    bar_row = dbc.Row(
            [
                # Create the graduated bar to the left
                dbc.Col(
                    [
                        daq.GraduatedBar(
                            color={"gradient": True,
                                   "ranges": {app_color['green']: [0, 60],
                                              app_color['yellow']: [60, 75],
                                              app_color['red']: [75, 100]}
                                   },
                            showCurrentValue=False,
                            label=str(name),
                            id=str(name) + "-bar",
                            max=100,
                            step=2,
                            style={"padding-bottom": "5px"},
                            size=205
                        )
                    ]
                ),
                # Create the text displaying the actual data on the right
                dbc.Col(
                    [
                        html.P(
                            id=str(name) + "-bar-text",
                            children=['No data'],
                            style={'position': 'absolute',
                                   'bottom': '0',
                                   'font-size': '.9rem'},
                        )
                    ],
                )
            ]
    )
    # Define the graduated bar styling and return it
    return bar_row

def update_liveplot_helper(trace_dict, dropdown_value):
    """Helper function to populate liveplot depending on selected pollutants"""

    # Find number of pollutants selected to be plotted
    numplots = len(list(dropdown_value))

    # Create figure layout depending on number of selected pollutants. Also create a list of figure titles.
    if numplots == 1:
        figure_title_list = [dropdown_value[0]]
        figspec = [
            [{'rowspan': 2, 'colspan': 2}, None],
            [None, None]
        ]
    elif numplots == 2:
        figure_title_list = dropdown_value[:2]
        figspec = [
            [{'rowspan': 2}, {'rowspan': 2}],
            [None, None]
        ]
    elif numplots == 3:
        figure_title_list = dropdown_value[:3]
        figspec = [
            [{}, {}],
            [{'colspan': 2}, None]
        ]
    elif numplots == 4:
        figure_title_list = dropdown_value[:4]
        figspec = [
            [{}, {}],
            [{}, {}]
        ]
    else:
        figure_title_list = ["Select pollutants"]
        figspec = [
            [None, None],
            [None, None]
        ]

    # Create the figure container with row and column divisions.
    fig = make_subplots(rows=2, cols=2, specs=figspec, subplot_titles=figure_title_list, vertical_spacing=0.10)

    # Populate the divisions with scatter plots depending on number of selected pollutants.
    if numplots > 0:
        # Add a scatter in its respective row and column using the dropdown values to select pollutants
        fig.add_scatter(y=list(trace_dict[list(dropdown_value)[0]]['y']),
                        x=list(trace_dict[list(dropdown_value)[0]]['x']),
                        row=1, col=1, name=dropdown_value[0], mode='lines')

        # Update axis labels to correspond with those being plotted using the label dictionary.
        fig.update_xaxes(title_text="Time (HH:MM:SS)", row=1, col=1),
        fig.update_yaxes(title_text=labeldict[dropdown_value[0]], row=1, col=1, range=y_range_dict[dropdown_value[0]])

    if numplots > 1:
        # Add a scatter in its respective row and column using the dropdown values to select pollutants
        fig.add_scatter(y=list(trace_dict[list(dropdown_value)[1]]['y']),
                        x=list(trace_dict[list(dropdown_value)[1]]['x']),
                        row=1, col=2, name=dropdown_value[1], mode='lines')

        # Update axis labels to correspond with those being plotted using the label dictionary.
        fig.update_xaxes(title_text="Time (HH:MM:SS)", row=1, col=2),
        fig.update_yaxes(title_text=labeldict[dropdown_value[1]], row=1, col=2, range=y_range_dict[dropdown_value[1]])

    if numplots > 2:
        # Add a scatter in its respective row and column using the dropdown values to select pollutants
        fig.add_scatter(y=list(trace_dict[list(dropdown_value)[2]]['y']),
                        x=list(trace_dict[list(dropdown_value)[2]]['x']),
                        row=2, col=1, name=dropdown_value[2], mode='lines')

        # Update axis labels to correspond with those being plotted using the label dictionary.
        fig.update_xaxes(title_text="Time (HH:MM:SS)", row=2, col=1),
        fig.update_yaxes(title_text=labeldict[dropdown_value[2]], row=2, col=1, range=y_range_dict[dropdown_value[2]])

    if numplots > 3:
        # Add a scatter in its respective row and column using the dropdown values to select pollutants
        fig.add_scatter(y=list(trace_dict[list(dropdown_value)[3]]['y']),
                        x=list(trace_dict[list(dropdown_value)[3]]['x']),
                        row=2, col=2, name=dropdown_value[3], mode='lines')

        # Update axis labels to correspond with those being plotted using the label dictionary.
        fig.update_xaxes(title_text="Time (HH:MM:SS)", row=2, col=2),
        fig.update_yaxes(title_text=labeldict[dropdown_value[3]], row=2, col=2, range=y_range_dict[dropdown_value[3]])

    # Apply plot styling
    fig.update_layout(
        autosize=True,
        margin=dict(t=20, b=20),
        plot_bgcolor=app_color["graph_bg"],
        showlegend=False,
    )

    # Set the number of ticks on the x axis in all plots to declutter the axis tick text.
    fig.update_xaxes(
        nticks=4
    )

    return fig

#define live plots
liveplot = dbc.Card(
    [
        dbc.Row(
            [
                dbc.Col(
                    html.H5("Pollutant Histogram"),
                ),
            ]
        ),
        dcc.Graph(
            id='liveplot',
            config={'displayModeBar': False},
            style={"height": "100%", "overflow": "hidden"}
        ),
        dbc.Row(
            [
                dbc.Col(
                    [
                        dcc.Dropdown(
                            id='graph-dropdown',
                            placeholder="Select plotted pollutants...",
                            options=instrumentdict,
                            multi=True,
                        ),
                    ],
                    width=6,
                ),
                # This col-row nest is ugly. It's the only way to output something pretty for an operator, so it stays.
                dbc.Col(
                    [
                        dbc.Row(
                            [
                                dbc.Col(
                                    [
                                        dbc.Input(
                                            id='mark-event-input',
                                            placeholder="Enter a short, descriptive event tag",
                                            className='appended-input'
                                        ),
                                        dbc.FormFeedback(
                                            "Event marker logged to file",
                                            valid=True
                                        ),
                                        dbc.FormFeedback(
                                            "Error marking event",
                                            valid=False
                                        )
                                    ]
                                ),
                                dbc.Col(
                                    [
                                        dbc.Button(
                                            "Mark Event",
                                            id="mark-event",
                                            n_clicks=0,
                                            size="sm",
                                            color="warning",
                                            className="appended-button"
                                        ),
                                    ],
                                    width="auto"
                                )
                            ],
                        )
                    ],
                    width=6
                )
            ]
        ),
    ],
    style={"padding": "5px 5px 5px 5px"},
    className="h-100",
)

#define graduated bars on right
livebar = dbc.Card(
    [
        html.H5("Live Pollutant Data"),
        create_graduatedbar_helper("NO2"),
        create_graduatedbar_helper("WCPC"),
        create_graduatedbar_helper("O3"),
        create_graduatedbar_helper("CO"),
        create_graduatedbar_helper("CO2"),
        create_graduatedbar_helper("NO"),
        create_graduatedbar_helper("WS"),
        create_graduatedbar_helper("WD"),
    ],
    style={"padding": "5px 5px 5px 5px"},
    className="h-100"
)

#define wind direction plot
wind_direction = dbc.Card(
    [
        html.H5(
            "Wind Direction", className="graph__title"
        ),
        dcc.Graph(
            id="wind-direction",
            config={'displayModeBar': False},
            figure=dict(
                layout=dict(
                   plot_bgcolor=app_color["graph_bg"],
                    paper_bgcolor=app_color["graph_bg"],
                )
            ),
            style={"width": "100%",
                   "height": "100%",
                   "justify-content": "center",
                   "align-items": "center",
                   "display": "flex",
                   }
        ),
    ],
    style={"padding": "5px 5px 5px 5px"},
    className="h-100"
)

#define the layout of the dashboard
app.layout = dbc.Container(
    [
        # Data Display Panel
        dbc.Row(
            [
                # Left column
                dbc.Col(
                    html.Div(
                        liveplot,
                        style={"height": "96vh"}
                    ),
                    width=9,

                ),
                # Right column
                dbc.Col(
                    [
                        # Top
                        html.Div(
                            livebar,
                            style={"height": "52vh"}
                        ),
                        # Bottom
                        html.Div(
                            wind_direction,
                            style={"height": "44vh"}
                        ),
                    ],
                    width=3,
                ),
            ],
        ),
        # Update interval for the wind direction module
        dcc.Interval(
            id='wind-interval',
            interval=interval_s * 1000,
            n_intervals=0,
        ),
        # Update interval for instrument input
        dcc.Interval(
            id='daq-interval',
            interval=1000,
            n_intervals=0
        ),
        # Update interval for the live figures
        dcc.Interval(
            id='figure-interval',
            interval=500,
            n_intervals=0
        ),
        html.Div(
            id='dump-hdiv',
            style={'display': 'none'}
        )
    ],
    fluid=True,
)

avs=0


'''########################################
## SECTION 4. CSV data marking functions ##
########################################'''
avs=0
#shared auto event marking function
def auto_event_mark(auto_event_name,algorithm,pollutant):
    try:
        # get today's date
        this_day = dt.date.today()
        # Create a file name.
        filename = log_folder_path + "/Event Markers " + str(this_day) + ".csv"
        txt_filename = log_folder_path + "/Event Markers Backup " + str(this_day) + ".txt"

        # Check if the file exists already.
        file_exists = os.path.isfile(filename)
        with open(filename, 'a', newline='\n') as file, open(txt_filename, 'a', newline='\n') as txt_file:
            # Create a dictionary with our seven csv columns
            markerdict = dict.fromkeys(['Type','Pollutant','Event Tag', 'Time','NO2 (ppb)', 'WCPC (#/cm^3)', 'O3 (ppb)', 'CO (ppm)', 'CO2 (ppm)','NO (ppb)','WS (m/s)','WD (degrees)'])

            # Populate our columns with the user input and current time and current values of pollutants
            markerdict['Type'] = algorithm
            markerdict['Pollutant'] = pollutant
            markerdict['Time'] = (dt.datetime.now()).strftime("%Y-%m-%d %H:%M:%S")
            markerdict['Event Tag'] = auto_event_name
            markerdict['NO2 (ppb)'] = no2_trace_y[-1]
            markerdict['WCPC (#/cm^3)'] = wcpc_trace_y[-1]
            markerdict['O3 (ppb)'] = o3_trace_y[-1]
            markerdict['CO (ppm)'] = co_trace_y[-1]
            markerdict['CO2 (ppm)'] = co2_trace_y[-1]
            markerdict['NO (ppb)'] = no_trace_y[-1]
            markerdict['WS (m/s)'] = ws_trace_y[-1]
            markerdict['WD (degrees)'] = wd_trace_y[-1]

            # Write our data to our csv
            writer = csv.DictWriter(file, delimiter=',', fieldnames=list(markerdict.keys()))

            # prepare a string to be written to our txt and then write to it
            txt_string = str((dt.datetime.now()).strftime("%Y-%m-%d %H:%M:%S")) + ", " + auto_event_name + ", " + str(
                no2_trace_y[-1]) + ", " + str(wcpc_trace_y[-1]) + ", " + str(o3_trace_y[-1]) + ", " + str(
                co_trace_y[-1]) + ", " + str(co2_trace_y[-1]) +', ' +str(no_trace_y[-1]) +', ' +str(ws_trace_y[-1]) +', ' +str(wd_trace_y[-1]) +"\n"
            txt_file.write(txt_string)

            # If the log file did not already exist, write the column headers.
            if not file_exists:
                writer.writeheader()

            # Write our row of data
            writer.writerow(markerdict)
    # If an IO error occurs, do the following
    except IOError:
        print(
            "Error marking event. Check log folder read/write permissions or run bash script as administrator")
        return False, not False

#sensor transcript function, writes current row to sensor transcript
def sensor_dump():
    global index_clock
    try:
        # get today's date
        this_day = dt.date.today()
        # Create a file name.
        filename = log_folder_path + "/Sensor Transcript " + str(this_day) + ".csv"
        txt_filename = log_folder_path + "/Sensor Transcript Backup " + str(this_day) + ".txt"

        # Check if the file exists already.
        file_exists = os.path.isfile(filename)
        with open(filename, 'a', newline='\n') as file, open(txt_filename, 'a', newline='\n') as txt_file:
            # Create a dictionary with our seven csv columns
            markerdict = dict.fromkeys(['Row','Time','NO2 (ppb)', 'WCPC (#/cm^3)', 'O3 (ppb)', 'CO (ppm)', 'CO2 (ppm)','NO (ppb)','WS (m/s)','WD (degrees)'])

            # Populate our columns with the user input and current time and current values of pollutants
            markerdict['Row'] = index_clock
            markerdict['Time'] = (dt.datetime.now()).strftime("%Y-%m-%d %H:%M:%S")
            markerdict['NO2 (ppb)'] = no2_trace_y[-1]
            markerdict['WCPC (#/cm^3)'] = wcpc_trace_y[-1]
            markerdict['O3 (ppb)'] = o3_trace_y[-1]
            markerdict['CO (ppm)'] = co_trace_y[-1]
            markerdict['CO2 (ppm)'] = co2_trace_y[-1]
            markerdict['NO (ppb)'] = no_trace_y[-1]
            markerdict['WS (m/s)'] = ws_trace_y[-1]
            markerdict['WD (degrees)'] = wd_trace_y[-1]

            # Write our data to our csv
            writer = csv.DictWriter(file, delimiter=',', fieldnames=list(markerdict.keys()))

            # prepare a string to be written to our txt and then write to it
            txt_string = str(index_clock)+", "+str((dt.datetime.now()).strftime("%Y-%m-%d %H:%M:%S")) + ", " +str(
                no2_trace_y[-1]) + ", " + str(wcpc_trace_y[-1]) + ", " + str(o3_trace_y[-1]) + ", " +str(
                co_trace_y[-1]) + ", " + str(co2_trace_y[-1]) + ", "+ str(no_trace_y[-1])  +', ' +str(ws_trace_y[-1]) +', ' +str(wd_trace_y[-1]) +"\n"
            txt_file.write(txt_string)
            index_clock += 1

            # If the log file did not already exist, write the column headers.
            if not file_exists:
                writer.writeheader()

            # Write our row of data
            writer.writerow(markerdict)
    # If an IO error occurs, do the following
    except IOError:
        print(
            "Error marking event. Check log folder read/write permissions or run bash script as administrator")
        return False, not False

#function for reading commands and executing them
def read_command(raw_command):
    #exiting function if not a command
    global command_character
    if raw_command[0] != command_character:
        return None
    #setting all to lowercase and removing spaces
    command=""
    raw_command = raw_command.lower()
    for p in raw_command:
        if p != " ":
            command += p
    #marking as a valid command for tagging purposes
    global is_valid_command
    is_valid_command = True

    ##################
    ## COMMAND LIST ##
    ##################

    #change of A1_coeff command
    """
    '*coeff = 12 - no2' #changes A1_coeff for no2 to 12
    '*coeff = 456 - wcpc' #changes A1_coeff for wcpc to 456
    """
    if command[1:7] == "coeff=":
        global A1_coeff

        #finding new coeff and pollutant from command
        new_coeff=""
        pollutant=""
        for i in range(7,len(command)):
            if command[i] == "-":
                pollutant = command[(i+1):len(command)]
                break
            new_coeff += command[i]

        print("changing A1_coeff for " + pollutant + " to " + new_coeff)
        A1_coeff[pollutant] = int(new_coeff)
        return None

    #change of A1_percentile command
    """
    '*percentile = 75 - no2' #changes A1_percentile for no2 to 75
    '*percentile = 50 - wcpc' #changes A1_coeff for wcpc to 50
    """
    if command[1:12] == "percentile=":
        global A1_percentile

        # finding new percentile and pollutant from command
        new_percentile = ""
        pollutant = ""
        for i in range(12, len(command)):
            if command[i] == "-":
                pollutant = command[(i + 1):len(command)]
                break
            new_percentile += command[i]

        print("changing A1_percentile for " + pollutant + " to " + new_percentile)
        A1_percentile[pollutant] = int(new_percentile)
        return None

    #change of AQ_thresh command
    """
    '*AQ_thresh = 40 - no2' #changes AQ_thresh for no2 to 40
    '*AQ_thresh = 8000 - wcpc' #changes AQ_thresh for wcpc to 8000
     """
    if command[1:11] == "aq_thresh=":
        global AQ_thresh

        # finding new thresh and pollutant from command
        new_thresh = ""
        pollutant = ""
        for i in range(11, len(command)):
            if command[i] == "-":
                pollutant = command[(i + 1):len(command)]
                break
            new_thresh += command[i]

        print("changing AQ_thresh for " + pollutant + " to " + new_thresh)
        AQ_thresh[pollutant] = int(new_thresh)
        return None

    #simple A1, AQ, and wind direction alert toggle switch
    """
    '*tA1' - toggles A1 on/off
    '*tAQ' - toggles AQ on/off
    '*tWDRW' - toggles wind direction alert on/off
    """
    if command[1] == 't':
        global toggle_type
        if command[2:4] == 'a1':
            global A1
            if A1 == True:
                A1 = False
                print("toggling A1 OFF")
                toggle_type = '(A1 OFF)'
                return None
            else:
                A1 = True
                print("toggling A1 ON")
                toggle_type = '(A1 ON)'
                return None

        if command[2:4] == 'aq':
            global AQ
            if AQ == True:
                AQ = False
                print("toggling AQ OFF")
                toggle_type = '(AQ OFF)'
                return None
            else:
                AQ = True
                print("toggling AQ ON")
                toggle_type = '(AQ ON)'
                return None

        if command[2:5] == 'wdrw':
            global enable_wind_direction_alert
            if enable_wind_direction_alert:
                enable_wind_direction_alert = False
                print("toggling wind direction range warning OFF")
                toggle_type = '(WDRW OFF)'
                return None
            else:
                enable_wind_direction_alert = True
                print("toggling wind direction range warning ON")
                toggle_type = '(WDRW ON)'
                return None



    #change of graph y axes range command (if statement accounting for both possible notations)
    """
    '*y_range = 0,80-no2' #changes graph y_range for no2 to [0,80]
    '*y_range = 6000,9000 - wcpc' #changes graph y_range for wcpc to [6000,9000]
    OR (shortcut version)
    '*0,80 no2' #changes graph y_range for no2 to [0,80]
    '*6000,9000 wcpc' #changes graph y_range for wcpc to [6000,9000]
    """
    if (command[1:9] == "y_range=") or (command[1].isnumeric()):
        global y_range_dict

        #alternate shortcut version
        if command[1].isnumeric():
            lower_bound = ""
            upper_bound = ""
            pollutant = ""
            for i in range(1,len(command)):
                if command[i] == ",":
                    for f in range((i+1),len(command)):
                        if command[f].isnumeric() == False:
                            pollutant = command[f:len(command)]
                            break
                        upper_bound += command[f]
                    break
                lower_bound += command[i]

            print("changing graph y_range for " + pollutant + " to [" + lower_bound + "," + upper_bound + "]")
            y_range_dict[pollutant.upper()] = [int(lower_bound), int(upper_bound)]
            return None


        #finding interval lower bound, upper bound, and pollutant from command
        lower_bound=""
        upper_bound=""
        pollutant = ""
        for i in range(9,len(command)):
            if command[i] == ",":
                for f in range((i+1),len(command)):
                    if command[f] == "-":
                        pollutant = command[(f+1):len(command)]
                        break
                    upper_bound += command[f]
                break
            lower_bound += command[i]


        print("changing graph y_range for " + pollutant + " to [" + lower_bound +","+upper_bound+"]" )
        y_range_dict[pollutant.upper()] = [int(lower_bound),int(upper_bound)]
        return None



    #autoscale toggle switch
    """
    '*AS no2' - toggles autoscale on/off for NO2
    '*AS wcpc' - toggles autoscale on/off for NO2
    """
    if command[1:3] == "as":
        global enable_autoscale_dict
        pollutant = ''

        for i in range(3, len(command)):
            pollutant += command[i]

        pollutant = pollutant.upper()

        if enable_autoscale_dict[pollutant]:
            print("Disabling autoscale for "+pollutant)
            enable_autoscale_dict[pollutant] = False
            y_range_dict[pollutant] = y_range_dict_original[pollutant]
        else:
            print("Enabling autoscale for " + pollutant)
            enable_autoscale_dict[pollutant] = True


    #change wind direction alert range
    """
    '*WDrange = 150,180' - changes wind direction alert range to 150 - 180
    '*WDrange = 200,300' - changes wind direction alert range to 200 - 300
    """
    if command[1:9] == 'wdrange=':
        global wind_direction_alert_range

        # finding interval lower bound, upper bound, and pollutant from command
        lower_bound = ""
        upper_bound = ""
        for i in range(9, len(command)):
            if command[i] == ",":
                for f in range((i + 1), len(command)):
                    upper_bound += command[f]
                break
            lower_bound += command[i]

        print("changing wind direction alert range to [" + lower_bound + "," + upper_bound + "]")
        wind_direction_alert_range = [int(lower_bound), int(upper_bound)]
        return None


    ##

#mark event function for manually marking events. This function is also used to call the read_command function. The
#is_valid_command variable is fed through both this and the read_command function
@app.callback([Output('mark-event-input', 'valid'),
               Output('mark-event-input', 'invalid')],
              Input('mark-event', 'n_clicks'),
              State('mark-event-input', 'value'))
def mark_event(n, eventtag):
    """Marks an event with the current time and a user input. Handles exceptions and returns user feedback"""

    # Ensure the button has been pressed
    if n != 0:
        if eventtag:
            # try except else statements allow us to handle exceptions due to read/write errors
            try:

                ################################################################################
                read_command(eventtag)
                global is_valid_command
                global toggle_type

                '''
                #special commands for NO2 A1 settings and AQ setting
                if eventtag[0] == "%":
                    global A1_coeff
                    global A1_percentile

                    #change of A1_coeff for no2
                    new_coeff = ""
                    if eventtag[1].lower() == "c":
                        for i in range(2, len(eventtag)):
                            new_coeff += eventtag[i]
                        print("changing A1_coeff for NO2 to "+new_coeff)
                        A1_coeff["no2"] = int(new_coeff)

                    #change of A1_percentile for no2
                    new_percentile = ""
                    if eventtag[1].lower() == "p":
                        for i in range(2, len(eventtag)):
                            new_percentile += eventtag[i]
                        print("changing A1_percentile for NO2 to " + new_percentile)
                        A1_percentile["no2"] = int(new_percentile)

                    #change AQ thresh for no2
                    new_thresh = ""
                    if eventtag[1].lower() == "a":
                        for i in range(2, len(eventtag)):
                            new_thresh += eventtag[i]
                        print("changing AQ_thresh for NO2 to " + new_thresh)
                        AQ_thresh["no2"] = int(new_thresh)
                '''
                ###############################################################################

                #get today's date
                this_day = dt.date.today()
                # Create a file name.
                filename = log_folder_path + "/Event Markers "+str(this_day)+".csv"
                txt_filename = log_folder_path + "/Event Markers Backup " + str(this_day) + ".txt"

                # Check if the file exists already.
                file_exists = os.path.isfile(filename)
                with open(filename, 'a', newline='\n') as file, open(txt_filename, 'a',newline='\n') as txt_file:
                    # Create a dictionary with our seven csv columns
                    markerdict = dict.fromkeys(['Type','Pollutant','Event Tag', 'Time','NO2 (ppb)', 'WCPC (#/cm^3)', 'O3 (ppb)', 'CO (ppm)', 'CO2 (ppm)','NO (ppb)','WS (m/s)','WD (degrees)'])


                    # Populate our columns with the user input and current time and current values of pollutants
                    #determining if event should be marked as a manual event or a command
                    if is_valid_command:
                        markerdict['Type'] = "command"
                    else:
                        markerdict['Type'] = "manual"
                    #resetting the valid command switch
                    is_valid_command = False
                    markerdict['Pollutant'] = "-"
                    markerdict['Time'] = (dt.datetime.now()).strftime("%Y-%m-%d %H:%M:%S")

                    if toggle_type == ('(A1 OFF)' or '(A1 ON)' or '(AQ OFF)' or '(AQ ON)' or '(WDA ON)' or '(WDA OFF)'):
                        markerdict['Event Tag'] = str(eventtag)+' '+str(toggle_type)
                    else:
                        markerdict['Event Tag'] = eventtag
                    toggle_type = ''
                    markerdict['NO2 (ppb)'] = no2_trace_y[-1]
                    markerdict['WCPC (#/cm^3)'] = wcpc_trace_y[-1]
                    markerdict['O3 (ppb)'] = o3_trace_y[-1]
                    markerdict['CO (ppm)'] = co_trace_y[-1]
                    markerdict['CO2 (ppm)'] = co2_trace_y[-1]
                    markerdict['NO (ppb)'] = no_trace_y[-1]
                    markerdict['WS (m/s)'] = ws_trace_y[-1]
                    markerdict['WD (degrees)'] = wd_trace_y[-1]

                    # Write our data to our csv
                    writer = csv.DictWriter(file, delimiter=',', fieldnames=list(markerdict.keys()))

                    #prepare a string to be written to our txt and then write to it
                    txt_string = str((dt.datetime.now()).strftime("%Y-%m-%d %H:%M:%S"))+", "+str(eventtag)+", "+str(no2_trace_y[-1])+", "+str(wcpc_trace_y[-1])+", "+str(o3_trace_y[-1])+", "+str(co_trace_y[-1])+", "+str(co2_trace_y[-1])+', '+str(no_trace_y[-1])+', '+str(ws_trace_y[-1])+', '+str(wd_trace_y[-1])+"\n"
                    txt_file.write(txt_string)

                    # If the log file did not already exist, write the column headers.
                    if not file_exists:
                        writer.writeheader()

                    # Write our row of data
                    writer.writerow(markerdict)
            # If an IO error occurs, do the following
            except IOError:
                print(
                    "Error marking event. Check log folder read/write permissions or run bash script as administrator")
                return False, not False
            # If the try statement is successful, return a 'valid' response
            else:
                return True, not True
        # If the input field is blank, return an 'invalid' response
        else:
            return False, not False
    # If the button has not been pressed, return a 'not either' response
    else:
        return None, None

avs=0


'''################################
## SECTION 5. Get data functions ##
################################'''
abc=1
#these functions all use callbacks which is how we make the dashboard "live". They grab data from Redis

#helper function for autoscale
def compute_interval(input_trace,pollutant):
    global autoscale_padding_dict
    min_entry = min(input_trace)
    max_entry = max(input_trace)
    range = max_entry - min_entry
    padding_amount = range * ( autoscale_padding_dict[pollutant]/100 )

    '''
    if padding_amount<1:
        padding_amount = 1
    '''

    upper_bound = max_entry + padding_amount
    lower_bound = min_entry - padding_amount

    if lower_bound < 0:
        lower_bound = 0

    return [round(lower_bound,2), round(upper_bound,2)]

#get_no2_data function is ALSO responsible for calling our sensor_dump and zero_flush functions
@app.callback([Output('NO2-bar', 'value'),
               Output('NO2-bar-text', 'children')],
              Input('daq-interval', 'n_intervals'))
def get_no2_data(n):
    #set expected min and max for graduated bar
    minimumvalue = 0
    maximumvalue = 200

    #connect to redis
    conn = redis.Redis('localhost')

    #pulling data from redis
    redisdata = json.loads(conn.get('no2'))

    #prevent update if we grabbed duplicate data, otherwise we carry on
    if redisdata['time1'] == no2_trace_x[-1]:
        raise dash.exceptions.PreventUpdate
    else:
        #simulated data
        global simulated_or_real
        if simulated_or_real['no2'] == 'simulated':
            global simulated_data_filenames
            global no2_clock_x
            global no2_clock_y


            # grabbing simulated data
            if simulated_data_filetypes['no2'] == 'csv':
                no2_clock_y = no2_sim_csv[no2_clock_x-1]
                no2_clock_y = round(no2_clock_y, 2)
                no2_trace_y.append(no2_clock_y)
            else:
                xlsx_file = Path('SimData', simulated_data_filenames['no2'])
                wb_obj = openpyxl.load_workbook(xlsx_file)
                sheet = wb_obj.active
                no2_clock_y = sheet["B" + str(no2_clock_x)].value
                no2_clock_y = round(no2_clock_y, 2)
                no2_trace_y.append(no2_clock_y)
        else:
            no2_trace_y.append(round(redisdata['NO2'],2))

        no2_trace_x.append(redisdata['time1'])

        #dictionary of new value to be added, then append to container
        newvalue = dict(x=no2_trace_x, y=no2_trace_y)
        no2_trace_container.append(newvalue)


    #calling our algorithms
    A1ap(no2_trace_y, "no2")
    AQap(no2_trace_y, "no2")

    sensor_dump()
    zero_flush()


    ############ AUTOSCALE ###############
    if enable_autoscale_dict['NO2'] and len(no2_trace_y) > 3:
        global y_range_dict
        new_interval = compute_interval(no2_trace_y,'NO2')
        y_range_dict['NO2'] = new_interval
    ######################################

    if simulated_or_real['no2'] == 'simulated':
        no2_clock_x += 1
        return no2_clock_y, \
               str(no2_clock_y) + " " + labeldict['NO2'].split(' ')[1]
    else:
        return (redisdata['NO2'] - minimumvalue) / (maximumvalue - minimumvalue) * 100, \
            str(redisdata['NO2']) + " " + labeldict['NO2'].split(' ')[1]

@app.callback([Output('NO-bar', 'value'),
               Output('NO-bar-text', 'children')],
              Input('daq-interval', 'n_intervals'))
def get_no_data(n):
    #expected min and max for graduated bars
    minimumvalue = 0
    maximumvalue = 200

    #connect to redis
    conn = redis.Redis('localhost')

    #pull data from redis
    redisdata = json.loads(conn.get('no'))

    #prevent update if we grabbed duplicate data, otherwise we carry on
    if redisdata['time6'] == no_trace_x[-1]:
        raise dash.exceptions.PreventUpdate
    else:
        #simulated data
        global simulated_or_real
        if simulated_or_real['no'] == 'simulated':
            global simulated_data_filenames
            global no_clock_x
            global no_clock_y

            # grab simulated data
            if simulated_data_filetypes['no'] == 'csv':
                no_clock_y = no_sim_csv[no_clock_x-1]
                no_clock_y = round(no_clock_y, 2)
                no_trace_y.append(no_clock_y)
            else:
                xlsx_file = Path('SimData', simulated_data_filenames['no'])
                wb_obj = openpyxl.load_workbook(xlsx_file)
                sheet = wb_obj.active
                no_clock_y = sheet["B" + str(no_clock_x)].value
                no_clock_y = round(no_clock_y, 2)
                no_trace_y.append(no_clock_y)
        else:
            no_trace_y.append(round(redisdata['NO'],2))

        no_trace_x.append(redisdata['time6'])

        #create dictionary for new value to be plotted then append to container
        newvalue = dict(x=no_trace_x, y=no_trace_y)
        no_trace_container.append(newvalue)

        #call our algorithms
        A1ap(no_trace_y, "no")
        AQap(no_trace_y, "no")

        if enable_autoscale_dict['NO'] and len(no_trace_y) > 3:
            global y_range_dict
            new_interval = compute_interval(no_trace_y, 'NO')
            y_range_dict['NO'] = new_interval


        #return either real data or our simulated data
        if simulated_or_real['no'] == 'simulated':
            no_clock_x += 1
            return no_clock_y, \
                   str(no_clock_y) + " " + labeldict['NO'].split(' ')[1]
        else:
            return (redisdata['NO'] - minimumvalue) / (maximumvalue - minimumvalue) * 100, \
                str(redisdata['NO']) + " " + labeldict['NO'].split(' ')[1]

@app.callback([Output('WCPC-bar', 'value'),
               Output('WCPC-bar-text', 'children')],
              Input('daq-interval', 'n_intervals'))
def get_wcpc_data(n):
    #expected min and max for graduated bars
    minimumvalue = 1000
    maximumvalue = 20000

    #connect to redis
    conn = redis.Redis('localhost')

    #grab data from redis
    redisdata = json.loads(conn.get('wcpc'))

    #check if data is duplicate, otherwise carry on
    if redisdata['time2'] == wcpc_trace_x[-1]:
        raise dash.exceptions.PreventUpdate
    else:
        #simulated data
        global simulated_or_real
        if simulated_or_real['wcpc'] == 'simulated':
            global simulated_data_filenames
            global wcpc_clock_x
            global wcpc_clock_y

            #grabbing our simulated data
            if simulated_data_filetypes['wcpc'] == 'csv':
                wcpc_clock_y = wcpc_sim_csv[wcpc_clock_x-1]
                wcpc_clock_y = int(wcpc_clock_y)
                wcpc_trace_y.append(wcpc_clock_y)
            else:
                xlsx_file = Path('SimData', simulated_data_filenames['wcpc'])
                wb_obj = openpyxl.load_workbook(xlsx_file)
                sheet = wb_obj.active
                wcpc_clock_y = sheet["B" + str(wcpc_clock_x)].value
                wcpc_clock_y = int(wcpc_clock_y)
                wcpc_trace_y.append(wcpc_clock_y)
        else:
            wcpc_trace_y.append(int(redisdata['concentration']))

        wcpc_trace_x.append(redisdata['time2'])

        #create our new value dictionary, add to container, you know the drill at this point
        newvalue = dict(x=wcpc_trace_x, y=wcpc_trace_y)
        wcpc_trace_container.append(newvalue)

        #calling our algorithms
        A1ap(wcpc_trace_y, "wcpc")
        AQap(wcpc_trace_y, "wcpc")

        if enable_autoscale_dict['WCPC'] and len(wcpc_trace_y) > 3:
            global y_range_dict
            new_interval = compute_interval(wcpc_trace_y, 'WCPC')
            y_range_dict['WCPC'] = new_interval

        #return simulated data or real data
        if simulated_or_real['wcpc'] == 'simulated':
            wcpc_clock_x += 1
            return wcpc_clock_y, \
                   str(wcpc_clock_y) + " " + labeldict['WCPC'].split(' ')[1]
        else:
            return (redisdata['concentration'] - minimumvalue) / (maximumvalue - minimumvalue) * 100, \
                str(redisdata['concentration']) + " " + labeldict['WCPC'].split(' ')[1]

@app.callback([Output('O3-bar', 'value'),
               Output('O3-bar-text', 'children')],
              Input('daq-interval', 'n_intervals'))
def get_2b_data(n):
    #expected min and max for the graduated bar
    minimumvalue = 0
    maximumvalue = 100

    #connect to redis
    conn = redis.Redis('localhost')

    #grab data from redis
    redisdata = json.loads(conn.get('ozone'))

    #prevent update if we grabbed duplicate data, otherwise we carry on
    if redisdata['time3'] == o3_trace_x[-1]:
        raise dash.exceptions.PreventUpdate
    else:
        #simulated data
        global simulated_or_real
        if simulated_or_real['o3'] == 'simulated':
            global simulated_data_filenames
            global o3_clock_x
            global o3_clock_y


            if simulated_data_filetypes['o3'] == 'csv':
                o3_clock_y = o3_sim_csv[o3_clock_x-1]
                o3_clock_y = round(o3_clock_y, 2)
                o3_trace_y.append(o3_clock_y)
            else:
                xlsx_file = Path('SimData', simulated_data_filenames['o3'])
                wb_obj = openpyxl.load_workbook(xlsx_file)
                sheet = wb_obj.active
                o3_clock_y = sheet["B" + str(o3_clock_x)].value
                o3_clock_y = round(o3_clock_y, 2)
                o3_trace_y.append(o3_clock_y)
        else:
            o3_trace_y.append(round(redisdata['Ozone'],2))


        o3_trace_x.append(redisdata['time3'])


        newvalue = dict(x=o3_trace_x, y=o3_trace_y)
        o3_trace_container.append(newvalue)

        A1ap(o3_trace_y, "o3")
        AQap(o3_trace_y, "o3")

        if enable_autoscale_dict['O3'] and len(o3_trace_y) > 3:
            global y_range_dict
            new_interval = compute_interval(o3_trace_y, 'O3')
            y_range_dict['O3'] = new_interval


        if simulated_or_real['o3'] == 'simulated':
            o3_clock_x += 1
            return o3_clock_y, \
                   str(o3_clock_y) + " " + labeldict['O3'].split(' ')[1]
        else:
            return (redisdata['Ozone'] - minimumvalue) / (maximumvalue - minimumvalue) * 100, \
                str(redisdata['Ozone']) + " " + labeldict['O3'].split(' ')[1]

@app.callback([Output('CO-bar', 'value'),
               Output('CO-bar-text', 'children')],
              Input('daq-interval', 'n_intervals'))
def get_teledyne_CO_data(n):
    #min and max for graduated bar
    minimumvalue = 0
    maximumvalue = 20

    #connect to redis
    conn = redis.Redis('localhost')

    #pull data from redis
    redisdata = json.loads(conn.get('teledyne'))

    #check for duplicate
    if redisdata['time4'] == co_trace_x[-1]:
        raise dash.exceptions.PreventUpdate
    else:
        #simulated data
        global simulated_or_real
        if simulated_or_real['co'] == 'simulated':
            global simulated_data_filenames
            global co_clock_x
            global co_clock_y

            if simulated_data_filetypes['co'] == 'csv':
                co_clock_y = co_sim_csv[co_clock_x-1]
                co_clock_y = round(co_clock_y, 2)
                co_trace_y.append(co_clock_y)
            else:
                xlsx_file = Path('SimData', simulated_data_filenames['co'])
                wb_obj = openpyxl.load_workbook(xlsx_file)
                sheet = wb_obj.active
                co_clock_y = sheet["B" + str(co_clock_x)].value
                co_clock_y = round(co_clock_y, 2)
                co_trace_y.append(co_clock_y)
        else:
            co_trace_y.append(round(redisdata['CO'],2))

        co_trace_x.append(redisdata['time4'])

        newvalue = dict(x=co_trace_x, y=co_trace_y)
        co_trace_container.append(newvalue)

        A1ap(co_trace_y, "co")
        AQap(co_trace_y, "co")

        if enable_autoscale_dict['CO'] and len(co_trace_y) > 3:
            global y_range_dict
            new_interval = compute_interval(co_trace_y, 'CO')
            y_range_dict['CO'] = new_interval

        if simulated_or_real['co'] == 'simulated':
            co_clock_x += 1
            return co_clock_y, \
                   str(co_clock_y) + " " + labeldict['CO'].split(' ')[1]
        else:
            return (redisdata['CO'] - minimumvalue) / (maximumvalue - minimumvalue) * 100, \
               str(redisdata['CO']) + " " + labeldict['CO'].split(' ')[1]

@app.callback([Output('CO2-bar', 'value'),
               Output('CO2-bar-text', 'children')],
              Input('daq-interval', 'n_intervals'))
def get_licor_data(n):
    #expected min and max for graduated bar
    minimumvalue = 0
    maximumvalue = 1000

    #connect to redis and pull data
    conn = redis.Redis('localhost')
    redisdata = json.loads(conn.get('licor'))

    #check for duplicates
    if redisdata['time5'] == co2_trace_x[-1]:
        raise dash.exceptions.PreventUpdate
    else:
        #simulated data
        global simulated_or_real
        if simulated_or_real['co2'] == 'simulated':
            global simulated_data_filenames
            global co2_clock_x
            global co2_clock_y

            if simulated_data_filetypes['co2'] == 'csv':
                co2_clock_y = co2_sim_csv[co2_clock_x-1]
                co2_clock_y = round(co2_clock_y, 2)
                co2_trace_y.append(co2_clock_y)
            else:
                xlsx_file = Path('SimData', simulated_data_filenames['co2'])
                wb_obj = openpyxl.load_workbook(xlsx_file)
                sheet = wb_obj.active
                co2_clock_y = sheet["B" + str(co2_clock_x)].value
                co2_clock_y = round(co2_clock_y, 2)
                co2_trace_y.append(co2_clock_y)
        else:
            co2_trace_y.append(round(redisdata['CO2'],2))

        co2_trace_x.append(redisdata['time5'])

        newvalue = dict(x=co2_trace_x, y=co2_trace_y)
        co2_trace_container.append(newvalue)

        A1ap(co2_trace_y, "co2")
        AQap(co2_trace_y, "co2")

        if enable_autoscale_dict['CO2'] and len(co2_trace_y) > 3:
            global y_range_dict
            new_interval = compute_interval(co2_trace_y, 'CO2')
            y_range_dict['CO2'] = new_interval



        if simulated_or_real['co2'] == 'simulated':
            co2_clock_x += 1
            return co2_clock_y, \
                   str(co2_clock_y) + " " + labeldict['CO2'].split(' ')[1]
        else:
            return (redisdata['CO2'] - minimumvalue) / (maximumvalue - minimumvalue) * 100, \
               str(redisdata['CO2']) + " " + labeldict['CO2'].split(' ')[1]

@app.callback([Output('WS-bar', 'value'),
               Output('WS-bar-text', 'children')],
              Input('daq-interval', 'n_intervals'))
def get_wind_speed_data(n):
    #expected min and max for graduated bar
    minimumvalue = 0
    maximumvalue = 20

    #connect to redis and pull data
    conn = redis.Redis('localhost')
    redisdata = json.loads(conn.get('ws'))

    #check for duplicates
    if redisdata['time7'] == ws_trace_x[-1]:
        raise dash.exceptions.PreventUpdate
    else:
        #simulated data
        global simulated_or_real
        if simulated_or_real['ws'] == 'simulated':
            global simulated_data_filenames
            global ws_clock_x
            global ws_clock_y

            if simulated_data_filetypes['ws'] == 'csv':
                ws_clock_y = ws_sim_csv[ws_clock_x-1]
                ws_clock_y = round(ws_clock_y, 2)
                ws_trace_y.append(ws_clock_y)
            else:
                xlsx_file = Path('SimData', simulated_data_filenames['ws'])
                wb_obj = openpyxl.load_workbook(xlsx_file)
                sheet = wb_obj.active
                ws_clock_y = sheet["B" + str(ws_clock_x)].value
                ws_clock_y = round(ws_clock_y, 2)
                ws_trace_y.append(ws_clock_y)
        else:
            ws_trace_y.append(round(redisdata['WS'],2))


        ws_trace_x.append(redisdata['time7'])

        newvalue = dict(x=ws_trace_x, y=ws_trace_y)
        ws_trace_container.append(newvalue)

        A1ap(ws_trace_y, "ws")
        AQap(ws_trace_y, "ws")

        if enable_autoscale_dict['WS'] and len(ws_trace_y) > 3:
            global y_range_dict
            new_interval = compute_interval(ws_trace_y, 'WS')
            y_range_dict['WS'] = new_interval

        if simulated_or_real['ws'] == 'simulated':
            ws_clock_x += 1
            return ws_clock_y, \
                   str(ws_clock_y) + " " + labeldict['WS'].split(' ')[1]
        else:
            return (redisdata['WS'] - minimumvalue) / (maximumvalue - minimumvalue) * 100, \
               str(redisdata['WS']) + " " + labeldict['WS'].split(' ')[1]

@app.callback([Output('WD-bar', 'value'),
               Output('WD-bar-text', 'children')],
              Input('daq-interval', 'n_intervals'))
def get_wind_direction_data(n):
    #expected min and max for graduated bar
    minimumvalue = 0
    maximumvalue = 360

    #connect to redis and pull data
    conn = redis.Redis('localhost')
    redisdata = json.loads(conn.get('wd'))

    #check for duplicates
    if redisdata['time8'] == wd_trace_x[-1]:
        raise dash.exceptions.PreventUpdate
    else:
        #simulated data
        global simulated_or_real
        if simulated_or_real['wd'] == 'simulated':
            global simulated_data_filenames
            global wd_clock_x
            global wd_clock_y

            if simulated_data_filetypes['wd'] == 'csv':
                wd_clock_y = wd_sim_csv[wd_clock_x-1]
                wd_clock_y = round(wd_clock_y, 2)
                wd_trace_y.append(wd_clock_y)
            else:
                xlsx_file = Path('SimData', simulated_data_filenames['wd'])
                wb_obj = openpyxl.load_workbook(xlsx_file)
                sheet = wb_obj.active
                wd_clock_y = sheet["B" + str(wd_clock_x)].value
                wd_clock_y = round(wd_clock_y, 2)
                wd_trace_y.append(wd_clock_y)
        else:
            wd_trace_y.append(round(redisdata['WD'],2))

        wd_trace_x.append(redisdata['time8'])

        newvalue = dict(x=wd_trace_x, y=wd_trace_y)
        wd_trace_container.append(newvalue)

        A1ap(wd_trace_y, "wd")
        #AQap(wd_trace_y, "wd")
        wind_direction_alert(wd_trace_y, "wd")

        if enable_autoscale_dict['WD'] and len(wd_trace_y) > 3:
            global y_range_dict
            new_interval = compute_interval(wd_trace_y, 'WD')
            y_range_dict['WD'] = new_interval


        if simulated_or_real['wd'] == 'simulated':
            wd_clock_x += 1
            return wd_clock_y, \
                  str(wd_clock_y) + " " + labeldict['WD'].split(' ')[1]
        else:
            return (redisdata['WD'] - minimumvalue) / (maximumvalue - minimumvalue) * 100, \
            str(redisdata['WD']) + " " + labeldict['WD'].split(' ')[1]

@app.callback(Output('liveplot', 'figure'),
              Input('figure-interval', 'n_intervals'),
              State('graph-dropdown', 'value'))
def update_graph_scatter(n_intervals, dropdown_value):
    # Wrap the individual instrument dictionary containers into a dictionary with their corresponding keys.
    trace_dict = {
        'NO2': no2_trace_container[0],
        'WCPC': wcpc_trace_container[0],
        'O3': o3_trace_container[0],
        'CO': co_trace_container[0],
        'CO2': co2_trace_container[0],
        'NO': no_trace_container[0],
        'WS': ws_trace_container[0],
        'WD': wd_trace_container[0]
    }

    # Only update the figure if the user has selected a dropdown window
    if dropdown_value:
        fig = update_liveplot_helper(trace_dict, dropdown_value)
        return fig
    else:
        raise dash.exceptions.PreventUpdate

@app.callback(Output("wind-direction", "figure"),
               Input("wind-interval", "n_intervals"))
def gen_wind_direction(n):
#     """Generate the wind direction plot""

     # Get the current time and total time.
     now = dt.datetime.now()
     total_time = (now.hour * 3600) + (now.minute * 60) + now.second

     # Use the dataframe contained in the db folder.
     #df = get_wind_data_by_id(total_time)
     #val = df["Speed"].iloc[-1]
     #direction = [0, (df["Direction"][0] - 20), (df["Direction"][0] + 20), 0]

     conn = redis.Redis('localhost')
     redisdata = json.loads(conn.get('ws'))
     val = redisdata['WS']

     if simulated_or_real['ws'] == 'simulated':
         global ws_trace_y
         val = ws_trace_y[-1]

     redisdata = json.loads(conn.get('wd'))
     direction = [0,redisdata['WD']-20,redisdata['WD']+20,0]


     if simulated_or_real['wd'] == 'simulated':
         global wd_trace_y
         direction = [0,wd_trace_y[-1]-20,wd_trace_y[-1]+20,0]


     #direction = redisdata['WD']
     # Create polar traces.
     traces_scatterpolar = [
         {"r": [0, val, val, 0], "fillcolor": app_color['red']},
         {"r": [0, val * 0.65, val * 0.65, 0], "fillcolor": app_color['yellow']},
         {"r": [0, val * 0.3, val * 0.3, 0], "fillcolor": app_color['green']},
     ]

     # Create the data dictionary that will get passed onto the polar figure.
     data = [
         dict(
             type="scatterpolar",
             r=traces["r"],
             theta=direction,
             mode="lines",
             fill="toself",
             fillcolor=traces["fillcolor"],
             line={"color": "rgba(32, 32, 32, .6)", "width": 1},
         )
         for traces in traces_scatterpolar
     ]

     # Create the layout that styles the polar figure.
     layout = dict(
         margin=dict(t=10, b=10, l=10, r=10),
         responsive=True,
         font={"color": "black"},
         polar={
             "bgcolor": app_color["graph_bg"],
             #range = wind speed range
             "radialaxis": {"range": [0, 5], "angle": 45, "dtick": 2},
             "angularaxis": {"direction": "clockwise", "showline": True, "tickcolor": "black"},
         },
         showlegend=False,
     )

     # Return the data and layout to the polar figure.
     return dict(data=data, layout=layout)

# If the program is called as 'main' (e.g. not imported and ran from within another python script), do the following.


avs =0


'''#########################################
## SECTION 6. Settings and initialization ##
#########################################'''
if __name__ == '__main__':

    if exists('user_defined_settings.ini') == False:
        sys.exit("ERROR: \"user_defined_settings.ini\" config file not found, please run \"create_default_config.py\"")



    #config
    parser = ConfigParser(allow_no_value=True)
    parser.read('user_defined_settings.ini')
    #global counters, ONLY change when adding new pollutants, otherwise DON'T touch these
    avs =0
    A1_n = {
        "no2": 0,
        "wcpc": 0,
        "o3": 0,
        "co": 0,
        "co2": 0,
        "no": 0,
        "ws": 0,
        "wd": 0}
    A1_auto_event_count = {
        "no2": 1,
        "wcpc": 1,
        "o3": 1,
        "co": 1,
        "co2": 1,
        "no": 1,
        "ws": 1,
        "wd": 1
    }
    left_zero_popped = False
    is_valid_command = False
    toggle_type=''
    index_clock = 0
    AQ_auto_event_count = {
        "no2": 1,
        "wcpc": 1,
        "o3": 1,
        "co": 1,
        "co2": 1,
        "no": 1,
        "ws": 1,
        "wd": 1
    }
    AQ_over = {
    "no2": False,
    "wcpc": False,
    "o3": False,
    "co": False,
    "co2": False,
    "no": False,
    "ws": False,
    "wd": False}
    A2_n = {
        "no2": 0,
        "wcpc": 0,
        "o3": 0,
        "co": 0,
        "co2": 0,
        "no": 0,
        "ws": 0,
        "wd": 0}
    A2_auto_event_count = {
        "no2": 1,
        "wcpc": 1,
        "o3": 1,
        "co": 1,
        "co2": 1,
        "no": 1,
        "ws": 1,
        "wd": 1
    }
    avs =0

    ############################
    ## GLOBAL PARAM VARIABLES ##
    ############################

    #simple start
    enable_simple_start = parser.getboolean('simple_start','enable_simple_start')
    if enable_simple_start:
        if not pyuac.isUserAdmin():
            sys.exit('ERROR: Please run PyCharm as admin')
        redis_directory = parser.get('simple_start','redis_program_directory')
        os.system('cd '+redis_directory)
        os.system('start redis-server.exe')
        os.system('start redis-cli.exe')
        daq_script_name = parser.get('simple_start','DAQ_script_name')
        exec(open(daq_script_name).read())

    #graph y axes ranges
    y_range_dict = {
        "NO2": string_to_list_interval(parser.get('y-ranges','NO2')),
        "WCPC": string_to_list_interval(parser.get('y-ranges','WCPC')),
        "O3": string_to_list_interval(parser.get('y-ranges','O3')),
        "CO": string_to_list_interval(parser.get('y-ranges','CO')),
        "CO2": string_to_list_interval(parser.get('y-ranges','CO2')),
        "NO": string_to_list_interval(parser.get('y-ranges', 'NO')),
        "WS": string_to_list_interval(parser.get('y-ranges', 'WS')),
        "WD": string_to_list_interval(parser.get('y-ranges', 'WD'))
    }
    y_range_dict_original = {
        "NO2": string_to_list_interval(parser.get('y-ranges','NO2')),
        "WCPC": string_to_list_interval(parser.get('y-ranges','WCPC')),
        "O3": string_to_list_interval(parser.get('y-ranges','O3')),
        "CO": string_to_list_interval(parser.get('y-ranges','CO')),
        "CO2": string_to_list_interval(parser.get('y-ranges','CO2')),
        "NO": string_to_list_interval(parser.get('y-ranges', 'NO')),
        "WS": string_to_list_interval(parser.get('y-ranges', 'WS')),
        "WD": string_to_list_interval(parser.get('y-ranges', 'WD'))
    }
    enable_autoscale_dict = {
        "NO2": parser.getboolean('y-ranges','as_NO2'),
        "WCPC": parser.getboolean('y-ranges','as_WCPC'),
        "O3": parser.getboolean('y-ranges','as_O3'),
        "CO": parser.getboolean('y-ranges','as_CO'),
        "CO2": parser.getboolean('y-ranges','as_CO2'),
        "NO": parser.getboolean('y-ranges', 'as_NO'),
        "WS": parser.getboolean('y-ranges', 'as_WS'),
        "WD": parser.getboolean('y-ranges', 'as_WD')
    }

    #not using these settings, they're just unnecessary and confusing
    '''
    autoscale_padding_dict = {
        "NO2": parser.getint('y-ranges','autoscale_padding_percentage_NO2'),
        "WCPC": parser.getint('y-ranges','autoscale_padding_percentage_WCPC'),
        "O3": parser.getint('y-ranges','autoscale_padding_percentage_O3'),
        "CO": parser.getint('y-ranges','autoscale_padding_percentage_CO'),
        "CO2": parser.getint('y-ranges','autoscale_padding_percentage_CO2'),
        "NO": parser.getint('y-ranges', 'autoscale_padding_percentage_NO'),
        "WS": parser.getint('y-ranges', 'autoscale_padding_percentage_WS'),
        "WD": parser.getint('y-ranges', 'autoscale_padding_percentage_WD')
    }
    '''
    autoscale_padding_dict = {
        "NO2": 10,
        "WCPC": 10,
        "O3": 10,
        "CO": 10,
        "CO2": 10,
        "NO": 10,
        "WS": 10,
        "WD": 10
    }

    #log folder path
    log_folder_path = parser.get('log_directory','log_files_path')
    if log_folder_path == '':
        sys.exit("ERROR: Please specify a directory in the [log_directory] \"log_files_path\" setting")
    if '\\' in log_folder_path:
        log_folder_path.replace("\\", "/")
    if log_folder_path.endswith('/'):
        log_folder_path.removesuffix('/')

    #command character
    #command_character = parser.get('command_char','command_character')
    command_character = "*" #manual override

    #Algorithim circuit breaker
    A1 = parser.getboolean('algorithm_circuit_breaker','A1_on')
    #A2 = parser.getboolean('algorithm_circuit_breaker','A2_on')
    A2 = False
    AQ = parser.getboolean('algorithm_circuit_breaker','AQ_on')

    #wind direction alert settings
    enable_wind_direction_alert = parser.getboolean('wind_direction_range_warning','enable_wdrw')
    wind_direction_alert_range = string_to_list_interval(parser.get('wind_direction_range_warning','range'))

    #A1 settings
    A1_coeff = {
    "no2": parser.getint('A1_coeff','NO2'),
    "wcpc": parser.getint('A1_coeff','WCPC'),
    "o3": parser.getint('A1_coeff','O3'),
    "co": parser.getint('A1_coeff','CO'),
    "co2": parser.getint('A1_coeff','CO2'),
    "no": parser.getint('A1_coeff','NO'),
    "ws": parser.getint('A1_coeff','WS'),
    "wd": parser.getint('A1_coeff','WD') }
    A1_percentile = {
    "no2": parser.getint('A1_percentile','NO2'),
    "wcpc": parser.getint('A1_percentile','WCPC'),
    "o3": parser.getint('A1_percentile','O3'),
    "co": parser.getint('A1_percentile','CO'),
    "co2": parser.getint('A1_percentile','CO2'),
    "no": parser.getint('A1_percentile','NO'),
    "ws": parser.getint('A1_percentile','WS'),
    "wd": parser.getint('A1_percentile','WD')}
    A1_startup_bypass = parser.getint('A1_misc','startup_bypass')
    A1_thresh_bump_percentile = {
    "no2": parser.getint('A1_thresh_bump_percentile', 'NO2'),
    "wcpc": parser.getint('A1_thresh_bump_percentile', 'WCPC'),
    "o3": parser.getint('A1_thresh_bump_percentile', 'O3'),
    "co": parser.getint('A1_thresh_bump_percentile', 'CO'),
    "co2": parser.getint('A1_thresh_bump_percentile', 'CO2'),
    "no": parser.getint('A1_thresh_bump_percentile', 'NO'),
    "ws": parser.getint('A1_thresh_bump_percentile', 'WS'),
    "wd": parser.getint('A1_thresh_bump_percentile', 'WD')}

    #A2 settings
    A2_slope_thresh = {
    "no2": 0.4,
    "wcpc": 0.4,
    "o3": 0.4,
    "co": 0.4,
    "co2": 0.4,
    "no": 0.4,
    "ws": 0.4,
    "wd": 0.4
    }
    A2_hits_to_sink = {
    "no2": 3,
    "wcpc": 3,
    "o3": 3,
    "co": 3,
    "co2": 3,
    'no': 3,
    "ws": 3,
    "wd": 3
    }
    A2_interval = {
        "no2": deque([], maxlen=3),
        "wcpc": deque([], maxlen=3),
        "o3": deque([], maxlen=3),
        "co": deque([], maxlen=3),
        "co2": deque([], maxlen=3),
        "no": deque([], maxlen=3),
        "ws": deque([], maxlen=3),
        "wd": deque([], maxlen=3)
    }

    #AQ setting
    AQ_thresh = {
    "no2": parser.getint('AQ_thresh','NO2'),
    "wcpc": parser.getint('AQ_thresh','WCPC'),
    "o3": parser.getint('AQ_thresh','O3'),
    "co": parser.getint('AQ_thresh','CO'),
    "co2": parser.getint('AQ_thresh','CO2'),
    "no": parser.getint('AQ_thresh','NO'),
    "ws": parser.getint('AQ_thresh','WS')}

    #simulated or real switch
    simulated_or_real = {
        "no2": parser.get('real_or_simulated', 'NO2'),
        "wcpc": parser.get('real_or_simulated', 'WCPC'),
        "o3": parser.get('real_or_simulated', 'O3'),
        "co": parser.get('real_or_simulated', 'CO'),
        "co2": parser.get('real_or_simulated', 'CO2'),
        "no": parser.get('real_or_simulated', 'NO'),
        "ws": parser.get('real_or_simulated', 'WS'),
        "wd": parser.get('real_or_simulated', 'WD')
    }

    #check if using all real data
    all_real = True
    for i in simulated_or_real:
        if simulated_or_real[i] == 'simulated':
            all_real = False
            break

    #simulated or real switch error handling
    for i in simulated_or_real:
        if (simulated_or_real[i] != 'real') and (simulated_or_real[i] != 'simulated'):
            sys.exit('ERROR: the [real_or_simulated] \"'+i+'\" setting must be set to either \"simulated\" or \"real\"')

    #simulated data path and filenames
    simulated_data_path = parser.get('real_or_simulated', 'sim_data_path')

    if (simulated_data_path == '') and (all_real == False):
        sys.exit('ERROR: please specify a directory for the [real_or_simulated] \"sim_data_path\" setting')

    if '\\' in simulated_data_path:
        simulated_data_path.replace("\\", "/")

    if (simulated_data_path[-1] != '/'):
        simulated_data_path += '/'
    simulated_data_filenames = {
            "no2": '',
            "wcpc": '',
            "o3": '',
            "co": '',
            "co2": '',
            "no": '',
            "ws": '',
            "wd": ''
        }
    if simulated_or_real['no2'] == 'simulated':
        simulated_data_filenames['no2'] = simulated_data_path + parser.get('real_or_simulated','sim_NO2_filename')
    if simulated_or_real['wcpc'] == 'simulated':
        simulated_data_filenames['wcpc'] = simulated_data_path + parser.get('real_or_simulated','sim_WCPC_filename')
    if simulated_or_real['o3'] == 'simulated':
        simulated_data_filenames['o3'] = simulated_data_path + parser.get('real_or_simulated','sim_O3_filename')
    if simulated_or_real['co'] == 'simulated':
        simulated_data_filenames['co'] = simulated_data_path + parser.get('real_or_simulated','sim_CO_filename')
    if simulated_or_real['co2'] == 'simulated':
        simulated_data_filenames['co2'] = simulated_data_path + parser.get('real_or_simulated','sim_CO2_filename')
    if simulated_or_real['no'] == 'simulated':
        simulated_data_filenames['no'] = simulated_data_path + parser.get('real_or_simulated','sim_NO_filename')
    if simulated_or_real['ws'] == 'simulated':
        simulated_data_filenames['ws'] = simulated_data_path + parser.get('real_or_simulated','sim_WS_filename')
    if simulated_or_real['wd'] == 'simulated':
        simulated_data_filenames['wd'] = simulated_data_path + parser.get('real_or_simulated','sim_WD_filename')

    for i in simulated_or_real:
        if simulated_or_real[i] == 'simulated':
            if exists(simulated_data_filenames[i]) == False:
                sys.exit("ERROR: \""+simulated_data_filenames[i]+'\" file not found, please check [real_or_simlated] \"sim_'+i+'_filename\" setting')

    #simulated data file types
    simulated_data_filetypes = {
        "no2": '',
        "wcpc": '',
        "o3": '',
        "co": '',
        "co2": '',
        "no": '',
        "ws": '',
        "wd": ''
    }

    def get_extension(filename):
        last_dot_index = 0
        for i in range(0, len(filename)):
            if filename[i] == '.':
                last_dot_index = i
        return filename[(last_dot_index+1):len(filename)]

    for i in simulated_data_filetypes:
        simulated_data_filetypes[i] = get_extension(simulated_data_filenames[i])
        if ((simulated_data_filetypes[i] != 'xlsx') and (simulated_data_filetypes[i] != 'csv')) and (simulated_or_real[i] == False):
            sys.exit('ERROR: the simulated data file for ' + i + ' must be either a \".xlsx\" or \".csv\" file')

    #simulated data CSV handling
    if simulated_or_real['no2'] and simulated_data_filetypes['no2'] == 'csv':
        no2_sim_csv_df = pd.read_csv(simulated_data_filenames['no2'], names=['index','value'])
        no2_sim_csv = no2_sim_csv_df["value"].to_list()
    if simulated_or_real['wcpc'] and simulated_data_filetypes['wcpc'] == 'csv':
        wcpc_sim_csv_df = pd.read_csv(simulated_data_filenames['wcpc'], names=['index','value'])
        wcpc_sim_csv = wcpc_sim_csv_df["value"].to_list()
    if simulated_or_real['o3'] and simulated_data_filetypes['o3'] == 'csv':
        o3_sim_csv_df = pd.read_csv(simulated_data_filenames['o3'], names=['index','value'])
        o3_sim_csv = o3_sim_csv_df["value"].to_list()
    if simulated_or_real['co'] and simulated_data_filetypes['co'] == 'csv':
        co_sim_csv_df = pd.read_csv(simulated_data_filenames['co'], names=['index','value'])
        co_sim_csv = co_sim_csv_df["value"].to_list()
    if simulated_or_real['co2'] and simulated_data_filetypes['co2'] == 'csv':
        co2_sim_csv_df = pd.read_csv(simulated_data_filenames['co2'], names=['index','value'])
        co2_sim_csv = co2_sim_csv_df["value"].to_list()
    if simulated_or_real['no'] and simulated_data_filetypes['no'] == 'csv':
        no_sim_csv_df = pd.read_csv(simulated_data_filenames['no'], names=['index','value'])
        no_sim_csv = no_sim_csv_df["value"].to_list()
    if simulated_or_real['ws'] and simulated_data_filetypes['ws'] == 'csv':
        ws_sim_csv_df = pd.read_csv(simulated_data_filenames['ws'], names=['index','value'])
        ws_sim_csv = ws_sim_csv_df["value"].to_list()
    if simulated_or_real['wd'] and simulated_data_filetypes['wd'] == 'csv':
        wd_sim_csv_df = pd.read_csv(simulated_data_filenames['wd'], names=['index','value'])
        wd_sim_csv = wd_sim_csv_df["value"].to_list()


    #clock variables used for simulated data
    no2_clock_x = 1
    no2_clock_y = 0
    wcpc_clock_x = 1
    wcpc_clock_y = 0
    o3_clock_x = 1
    o3_clock_y = 0
    co_clock_x = 1
    co_clock_y = 0
    co2_clock_x = 1
    co2_clock_y = 0
    no_clock_x = 1
    no_clock_y = 0
    ws_clock_x = 1
    ws_clock_y = 0
    wd_clock_x = 1
    wd_clock_y = 0

    #run our server
    app.run_server(debug=True, dev_tools_ui=True, port=8090)
